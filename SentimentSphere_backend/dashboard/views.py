from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, Sum, Q, F
from django.http import JsonResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
import json
import logging
from datetime import timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from datetime import timedelta
from .models import FeedbackSource, FeedbackCategory, FeedbackItem, FeedbackBatch
from .serializers import (
    FeedbackSourceSerializer,
    FeedbackCategorySerializer,
    FeedbackItemSerializer,
    FeedbackBatchSerializer
)

logger = logging.getLogger(__name__)

@login_required
def dashboard_view(request):
    """
    Main dashboard view with enhanced summary statistics, trends, and visualizations
    """
    try:
        # Get basic counts with error handling
        total_feedback = FeedbackItem.objects.filter(client=request.user).count()
        processed_feedback = FeedbackItem.objects.filter(client=request.user, processed=True).count()
        total_batches = FeedbackBatch.objects.filter(client=request.user).count()
        
        # Calculate processing percentage safely
        if total_feedback > 0:
            processing_percentage = (processed_feedback / total_feedback) * 100
        else:
            processing_percentage = 0
        
        # Get sentiment distribution with prettier names
        sentiment_data = FeedbackItem.objects.filter(
            client=request.user, 
            processed=True,
            category__isnull=False
        ).values('category__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Get sources distribution
        source_data = FeedbackItem.objects.filter(
            client=request.user
        ).values('source__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Get recent feedback
        recent_feedback = FeedbackItem.objects.filter(
            client=request.user
        ).select_related('source', 'category').order_by('-created_at')[:10]
        
        # Calculate average ratings (only if ratings exist)
        avg_rating = FeedbackItem.objects.filter(
            client=request.user, 
            rating__isnull=False
        ).aggregate(avg=Avg('rating'))
        
        # Calculate average sentiment score
        avg_sentiment = FeedbackItem.objects.filter(
            client=request.user, 
            sentiment_score__isnull=False
        ).aggregate(avg=Avg('sentiment_score'))
        
        # Get counts for different rating levels
        rating_distribution = {}
        for i in range(1, 6):  # Ratings 1-5
            rating_distribution[i] = FeedbackItem.objects.filter(
                client=request.user, 
                rating=i
            ).count()
        
        # Get recent batches
        recent_batches = FeedbackBatch.objects.filter(
            client=request.user
        ).order_by('-upload_date')[:5]
        
        # Calculate trend data for last 7 days
        today = timezone.now().date()
        date_stats = []
        
        for i in range(6, -1, -1):
            target_date = today - timedelta(days=i)
            
            # Get counts for the day
            daily_stats = {
                'date': target_date.strftime('%b %d'),
                'positive': FeedbackItem.objects.filter(
                    client=request.user,
                    category__name='Positive',
                    created_at__date=target_date
                ).count(),
                'negative': FeedbackItem.objects.filter(
                    client=request.user,
                    category__name='Negative',
                    created_at__date=target_date
                ).count(),
                'neutral': FeedbackItem.objects.filter(
                    client=request.user,
                    category__name='Neutral',
                    created_at__date=target_date
                ).count()
            }
            
            date_stats.append(daily_stats)
        
        # Extract keywords from processed feedback (if available)
        # First check if any batches have keyword_data
        top_keywords = []
        try:
            batches_with_keywords = FeedbackBatch.objects.filter(
                client=request.user,
                keyword_data__isnull=False
            ).exclude(keyword_data='').order_by('-upload_date')[:3]
            
            # Combine keywords from recent batches
            for batch in batches_with_keywords:
                try:
                    batch_keywords = json.loads(batch.keyword_data)
                    # Add only if it's a valid list
                    if isinstance(batch_keywords, list):
                        top_keywords.extend(batch_keywords)
                except (json.JSONDecodeError, TypeError):
                    continue
            
            # Sort and get top keywords
            if top_keywords:
                # Sort by count (descending)
                top_keywords = sorted(top_keywords, key=lambda x: x.get('count', 0), reverse=True)[:16]
        except Exception as e:
            logger.error(f"Error processing keywords: {str(e)}")
            top_keywords = []
        
        # Calculate resolved issues and feature requests (example metrics)
        resolved_issues = {
            'total': FeedbackItem.objects.filter(
                client=request.user, 
                category__name='Bug Report'
            ).count(),
            'resolved': FeedbackItem.objects.filter(
                client=request.user, 
                category__name='Bug Report',
                processed=True
            ).count()
        }
        
        feature_requests = FeedbackItem.objects.filter(
            client=request.user, 
            category__name='Feature Request'
        ).count()
        
        # Calculate average response time (placeholder - customize based on your data model)
        # This assumes you have a response_time or similar field in your model
        avg_response_days = 2.4  # Placeholder value
        
        # Calculate customer satisfaction (based on ratings)
        if avg_rating['avg'] is not None:
            customer_satisfaction = (avg_rating['avg'] / 5) * 100
        else:
            customer_satisfaction = 0
            
        # Combine all data for the template
        context = {
            'total_feedback': total_feedback,
            'processed_feedback': processed_feedback,
            'processing_percentage': processing_percentage,
            'total_batches': total_batches,
            'sentiment_data': list(sentiment_data),
            'source_data': list(source_data),
            'recent_feedback': recent_feedback,
            'avg_rating': avg_rating['avg'] if avg_rating['avg'] is not None else 0,
            'avg_sentiment': avg_sentiment['avg'] if avg_sentiment['avg'] is not None else 0,
            'rating_distribution': rating_distribution,
            'recent_batches': recent_batches,
            'date_stats': date_stats,
            'top_keywords': top_keywords,
            'resolved_issues': resolved_issues,
            'feature_requests': feature_requests,
            'avg_response_days': avg_response_days,
            'customer_satisfaction': customer_satisfaction
        }
        
        return render(request, 'dashboard/dashboard.html', context)
    
    except Exception as e:
        logger.error(f"Dashboard error: {str(e)}")
        # Return a simplified dashboard with error messaging
        return render(request, 'dashboard/dashboard.html', {
            'error': True,
            'error_message': "There was an error loading the dashboard. Please try again later."
        })

@login_required
def feedback_list(request):
    """View to list all feedback items with filtering capabilities"""
    # Get all sources and categories for filtering
    sources = FeedbackSource.objects.all()
    categories = FeedbackCategory.objects.all()
    
    # Get query parameters for filtering
    source_id = request.GET.get('source')
    category_id = request.GET.get('category')
    rating = request.GET.get('rating')
    
    # Base queryset
    queryset = FeedbackItem.objects.filter(client=request.user)
    
    # Apply filters if provided
    filters_applied = False
    if source_id:
        queryset = queryset.filter(source_id=source_id)
        filters_applied = True
    if category_id:
        queryset = queryset.filter(category_id=category_id)
        filters_applied = True
    if rating:
        queryset = queryset.filter(rating=rating)
        filters_applied = True
    
    # Get feedback items and paginate
    feedback_items = queryset.select_related('source', 'category').order_by('-created_at')
    
    return render(request, 'dashboard/feedback_list.html', {
        'feedback_items': feedback_items,
        'sources': sources,
        'categories': categories,
        'filters_applied': filters_applied,
        'filtered_count': len(feedback_items) if filters_applied else None
    })

@login_required
def feedback_detail(request, pk):
    """View to show details of a single feedback item with enhanced context"""
    feedback = get_object_or_404(FeedbackItem, pk=pk, client=request.user)
    
    # Get the batch this feedback belongs to
    batch = feedback.batch
    
    # Get similar feedback based on category
    similar_feedback = []
    if feedback.category:
        similar_feedback = FeedbackItem.objects.filter(
            client=request.user,
            category=feedback.category
        ).exclude(id=feedback.id).order_by('-created_at')[:5]
    
    return render(request, 'dashboard/feedback_detail.html', {
        'feedback': feedback,
        'batch': batch,
        'similar_feedback': similar_feedback
    })

@login_required
def batch_list(request):
    """View to list all feedback batches with statistics"""
    # Get all batches with annotated statistics
    batches = FeedbackBatch.objects.filter(client=request.user).annotate(
        item_count=Count('feedback_items'),
        positive_count=Count('feedback_items', filter=Q(feedback_items__category__name='Positive')),
        negative_count=Count('feedback_items', filter=Q(feedback_items__category__name='Negative')),
        sentiment_avg=Avg('feedback_items__sentiment_score') 
    ).order_by('-upload_date')
    
    return render(request, 'dashboard/batch_list.html', {
        'batches': batches
    })

@login_required
def batch_detail(request, pk):
    """View to show details of a single batch with comprehensive statistics"""
    batch = get_object_or_404(FeedbackBatch, pk=pk, client=request.user)
    
    # Get feedback items for this batch
    feedback_items = FeedbackItem.objects.filter(
        client=request.user, 
        batch_id=batch.id
    ).select_related('category', 'source')
    
    # Get sentiment distribution
    sentiment_distribution = feedback_items.values('category__name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Get rating distribution
    rating_distribution = {}
    for i in range(1, 6):  # Ratings 1-5
        rating_distribution[i] = feedback_items.filter(rating=i).count()
    
    # Calculate average sentiment score
    avg_sentiment = feedback_items.aggregate(avg=Avg('sentiment_score'))
    
    # Extract keywords if available
    top_keywords = []
    if hasattr(batch, 'keyword_data') and batch.keyword_data:
        try:
            top_keywords = json.loads(batch.keyword_data)
        except json.JSONDecodeError:
            pass
    
    return render(request, 'dashboard/batch_detail.html', {
        'batch': batch,
        'feedback_items': feedback_items,
        'sentiment_distribution': sentiment_distribution,
        'rating_distribution': rating_distribution,
        'avg_sentiment': avg_sentiment['avg'] if avg_sentiment['avg'] is not None else 0,
        'top_keywords': top_keywords
    })

@login_required
def source_list(request):
    """View to list all feedback sources with statistics"""
    # Get all sources with statistics
    sources = FeedbackSource.objects.annotate(
        feedback_count=Count('feedback_items', filter=Q(feedback_items__client=request.user)),
        avg_sentiment=Avg('feedback_items__sentiment_score', filter=Q(feedback_items__client=request.user))
    )
    
    return render(request, 'dashboard/source_list.html', {'sources': sources})

@login_required
def category_list(request):
    """View to list all feedback categories with statistics"""
    # Get all categories with statistics
    categories = FeedbackCategory.objects.annotate(
        feedback_count=Count('feedback_items', filter=Q(feedback_items__client=request.user)),
        avg_sentiment=Avg('feedback_items__sentiment_score', filter=Q(feedback_items__client=request.user))
    )
    
    return render(request, 'dashboard/category_list.html', {'categories': categories})

# API Views with enhanced functionality and better error handling

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_feedback_list(request):
    """API endpoint to get all feedback items for the authenticated user with filters"""
    try:
        # Get query parameters for filtering
        source = request.query_params.get('source')
        category = request.query_params.get('category')
        processed = request.query_params.get('processed')
        rating = request.query_params.get('rating')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        # Base queryset
        queryset = FeedbackItem.objects.filter(client=request.user)
        
        # Apply filters if provided
        if source:
            queryset = queryset.filter(source_id=source)
        if category:
            queryset = queryset.filter(category_id=category)
        if processed:
            processed_bool = processed.lower() == 'true'
            queryset = queryset.filter(processed=processed_bool)
        if rating:
            queryset = queryset.filter(rating=rating)
        
        # Date range filtering
        if date_from:
            try:
                queryset = queryset.filter(created_at__gte=date_from)
            except ValueError:
                pass  # Ignore invalid date format
                
        if date_to:
            try:
                queryset = queryset.filter(created_at__lte=date_to)
            except ValueError:
                pass  # Ignore invalid date format
        
        # Order by creation date (newest first)
        queryset = queryset.order_by('-created_at')
        
        serializer = FeedbackItemSerializer(queryset, many=True)
        return Response(serializer.data)
    
    except Exception as e:
        logger.error(f"API error in feedback_list: {str(e)}")
        return Response(
            {"error": "An error occurred while processing your request."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def api_dashboard_stats(request):
    """API endpoint to get comprehensive dashboard statistics"""
    try:
        # Get basic counts
        total_feedback = FeedbackItem.objects.filter(client=request.user).count()
        processed_feedback = FeedbackItem.objects.filter(client=request.user, processed=True).count()
        total_batches = FeedbackBatch.objects.filter(client=request.user).count()
        
        # Get sentiment distribution
        sentiment_data = FeedbackItem.objects.filter(
            client=request.user, 
            processed=True,
            category__isnull=False
        ).values('category__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Get sources distribution
        source_data = FeedbackItem.objects.filter(
            client=request.user
        ).values('source__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Calculate trend data for last 7 days
        today = timezone.now().date()
        date_stats = []
        
        for i in range(6, -1, -1):
            target_date = today - timedelta(days=i)
            
            # Get counts for the day
            daily_stats = {
                'date': target_date.strftime('%b %d'),
                'positive': FeedbackItem.objects.filter(
                    client=request.user,
                    category__name='Positive',
                    created_at__date=target_date
                ).count(),
                'negative': FeedbackItem.objects.filter(
                    client=request.user,
                    category__name='Negative',
                    created_at__date=target_date
                ).count()
            }
            
            date_stats.append(daily_stats)
        
        # Calculate average rating and sentiment
        avg_stats = FeedbackItem.objects.filter(
            client=request.user
        ).aggregate(
            avg_rating=Avg('rating'),
            avg_sentiment=Avg('sentiment_score')
        )
        
        return Response({
            'total_feedback': total_feedback,
            'processed_feedback': processed_feedback,
            'total_batches': total_batches,
            'sentiment_distribution': list(sentiment_data),
            'source_distribution': list(source_data),
            'date_stats': date_stats,
            'avg_rating': avg_stats['avg_rating'] if avg_stats['avg_rating'] is not None else 0,
            'avg_sentiment': avg_stats['avg_sentiment'] if avg_stats['avg_sentiment'] is not None else 0
        })
        
    except Exception as e:
        logger.error(f"API error in dashboard_stats: {str(e)}")
        return Response(
            {"error": "An error occurred while processing your request."},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@login_required
def batch_delete(request, pk):
    """View to delete a feedback batch"""
    batch = get_object_or_404(FeedbackBatch, pk=pk, client=request.user)
    
    if request.method == 'POST':
        # Delete associated feedback items first
        FeedbackItem.objects.filter(batch=batch).delete()
        
        # Delete the batch
        batch.delete()
        
        messages.success(request, f'Batch "{batch.name}" deleted successfully.')
        return redirect('dashboard:batch_list')
    
    # If not POST, redirect to batch list
    return redirect('dashboard:batch_list')

@login_required
def feedback_update(request, pk):
    """Update a feedback item"""
    feedback = get_object_or_404(FeedbackItem, pk=pk, client=request.user)
    
    if request.method == 'POST':
        # Get form data
        category_id = request.POST.get('category')
        content = request.POST.get('content')
        rating = request.POST.get('rating')
        
        # Update the feedback item
        if category_id:
            try:
                category = FeedbackCategory.objects.get(pk=category_id)
                feedback.category = category
            except FeedbackCategory.DoesNotExist:
                pass
        
        feedback.content = content
        
        if rating:
            try:
                feedback.rating = int(rating)
            except ValueError:
                pass
        
        feedback.save()
        
        messages.success(request, 'Feedback updated successfully.')
    
    return redirect('dashboard:feedback_detail', pk=pk)

@login_required
def feedback_delete(request, pk):
    """Delete a feedback item"""
    feedback = get_object_or_404(FeedbackItem, pk=pk, client=request.user)
    
    if request.method == 'POST':
        # Store batch ID for redirection
        batch_id = feedback.batch.id if feedback.batch else None
        
        # Delete the feedback
        feedback.delete()
        
        messages.success(request, 'Feedback deleted successfully.')
        
        # Redirect to batch detail if from a batch, otherwise to feedback list
        if batch_id:
            return redirect('dashboard:batch_detail', pk=batch_id)
        else:
            return redirect('dashboard:feedback_list')
    
    # If not POST, redirect to detail page
    return redirect('dashboard:feedback_detail', pk=pk)

@login_required
def feedback_notes(request, pk):
    """Add or update notes for feedback"""
    feedback = get_object_or_404(FeedbackItem, pk=pk, client=request.user)
    
    if request.method == 'POST':
        # Get the notes from the form
        notes = request.POST.get('notes', '')
        
        # Check if the FeedbackItem model has a notes field
        if hasattr(feedback, 'notes'):
            # Update the feedback notes
            feedback.notes = notes
            feedback.save()
            
            messages.success(request, 'Notes updated successfully.')
        else:
            messages.error(request, 'Notes field not available.')
    
    return redirect('dashboard:feedback_detail', pk=pk)

@login_required
def feedback_process(request, pk):
    """Process an individual feedback item with AI"""
    feedback = get_object_or_404(FeedbackItem, pk=pk, client=request.user)
    
    if request.method == 'POST':
        # Get the prompt ID from the form
        prompt_id = request.POST.get('prompt_id')
        
        if not prompt_id:
            messages.error(request, 'Please select a prompt template.')
            return redirect('dashboard:feedback_detail', pk=pk)
        
        try:
            from analysis.models import AnalysisPrompt
            prompt = AnalysisPrompt.objects.get(pk=prompt_id, client=request.user)
            
            # Process the feedback using TextBlob
            try:
                from textblob import TextBlob
                
                # Use TextBlob for sentiment analysis
                blob = TextBlob(feedback.content)
                sentiment_score = blob.sentiment.polarity  # Range from -1 to 1
                
                # Determine category based on sentiment
                from dashboard.models import FeedbackCategory
                if "feature" in feedback.content.lower():
                    category, _ = FeedbackCategory.objects.get_or_create(name='Feature Request')
                elif "bug" in feedback.content.lower() or "issue" in feedback.content.lower():
                    category, _ = FeedbackCategory.objects.get_or_create(name='Bug Report')
                elif sentiment_score > 0.3:
                    category, _ = FeedbackCategory.objects.get_or_create(name='Positive')
                elif sentiment_score < -0.3:
                    category, _ = FeedbackCategory.objects.get_or_create(name='Negative')
                else:
                    category, _ = FeedbackCategory.objects.get_or_create(name='Neutral')
                
                # Update the feedback with the analysis results
                feedback.category = category
                feedback.sentiment_score = sentiment_score
                feedback.processed = True
                feedback.save()
                
                messages.success(request, 'Feedback processed successfully.')
                
            except Exception as e:
                messages.error(request, f'Error processing feedback: {str(e)}')
                
        except AnalysisPrompt.DoesNotExist:
            messages.error(request, 'Selected prompt template not found.')
    
    return redirect('dashboard:feedback_detail', pk=pk)

@login_required
def feedback_export(request, pk):
    """Export a single feedback item as JSON"""
    import json
    
    feedback = get_object_or_404(FeedbackItem, pk=pk, client=request.user)
    
    # Create a dictionary with feedback data
    data = {
        'id': feedback.id,
        'content': feedback.content,
        'category': feedback.category.name if feedback.category else None,
        'rating': feedback.rating,
        'sentiment_score': feedback.sentiment_score,
        'source': feedback.source.name,
        'processed': feedback.processed,
        'customer_id': feedback.customer_id,
        'feedback_date': feedback.feedback_date.isoformat() if feedback.feedback_date else None,
        'created_at': feedback.created_at.isoformat(),
    }
    
    # Add notes if available
    if hasattr(feedback, 'notes'):
        data['notes'] = feedback.notes
    
    # Create response with JSON data
    response = HttpResponse(json.dumps(data, indent=4), content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="feedback_{pk}.json"'
    
    return response

@login_required
def export_batches(request):
    """Export batches list as CSV, Excel, or JSON"""
    import json
    import csv
    from io import StringIO
    from django.db.models import Avg, Count, Q
    
    # Get format from request
    export_format = request.GET.get('format', 'csv')
    
    # Get selection option
    selection = request.GET.get('selection', 'all')
    
    # Determine which batches to include
    if selection == 'all':
        batches = FeedbackBatch.objects.filter(client=request.user)
    elif selection == 'filtered':
        # Get filter parameters from request
        source = request.GET.get('source')
        status = request.GET.get('status')
        date_range = request.GET.get('date_range')
        
        # Start with all batches for this user
        batches = FeedbackBatch.objects.filter(client=request.user)
        
        # Apply filters
        if source:
            batches = batches.filter(source_id=source)
        
        if status == 'processed':
            batches = batches.filter(processed=True)
        elif status == 'unprocessed':
            batches = batches.filter(processed=False)
        
        if date_range == 'today':
            today = timezone.now().date()
            batches = batches.filter(upload_date__date=today)
        elif date_range == 'week':
            start_of_week = timezone.now().date() - timedelta(days=timezone.now().weekday())
            batches = batches.filter(upload_date__date__gte=start_of_week)
        elif date_range == 'month':
            start_of_month = timezone.now().date().replace(day=1)
            batches = batches.filter(upload_date__date__gte=start_of_month)
        elif date_range == 'custom':
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            if date_from:
                batches = batches.filter(upload_date__date__gte=date_from)
            if date_to:
                batches = batches.filter(upload_date__date__lte=date_to)
    elif selection == 'selected':
        # Get selected batch IDs from request
        batch_ids = request.GET.getlist('batch_ids')
        batches = FeedbackBatch.objects.filter(client=request.user, id__in=batch_ids)
    else:
        batches = FeedbackBatch.objects.filter(client=request.user)
    
    # Determine what to include
    include_stats = request.GET.get('include_stats') == 'on'
    include_items = request.GET.get('include_items') == 'on'
    
    # Prepare data based on format
    if export_format == 'json':
        # Create JSON data
        data = []
        for batch in batches:
            batch_data = {
                'id': batch.id,
                'name': batch.name,
                'source': batch.source.name,
                'upload_date': batch.upload_date.isoformat(),
                'processed': batch.processed,
                'total_items': batch.total_items
            }
            
            if include_stats:
                # Add statistics
                feedback_items = FeedbackItem.objects.filter(batch=batch)
                batch_data['stats'] = {
                    'total_items': feedback_items.count(),
                    'processed_items': feedback_items.filter(processed=True).count(),
                    'avg_sentiment': feedback_items.aggregate(avg=Avg('sentiment_score'))['avg']
                }
                
                # Category counts
                categories = FeedbackCategory.objects.all()
                category_counts = {}
                for category in categories:
                    count = feedback_items.filter(category=category).count()
                    if count > 0:
                        category_counts[category.name] = count
                batch_data['category_counts'] = category_counts
            
            if include_items:
                # Add feedback items
                items_data = []
                feedback_items = FeedbackItem.objects.filter(batch=batch)
                for item in feedback_items:
                    items_data.append({
                        'id': item.id,
                        'content': item.content,
                        'category': item.category.name if item.category else None,
                        'rating': item.rating,
                        'sentiment_score': item.sentiment_score,
                        'processed': item.processed,
                        'customer_id': item.customer_id,
                        'created_at': item.created_at.isoformat()
                    })
                batch_data['items'] = items_data
            
            data.append(batch_data)
        
        # Create response with JSON data
        response = HttpResponse(json.dumps(data, indent=4), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="batches_export.json"'
        
    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Batches"
            
            # Add header row
            headers = ['ID', 'Name', 'Source', 'Upload Date', 'Processed', 'Total Items']
            if include_stats:
                headers.extend(['Avg Sentiment', 'Positive Count', 'Negative Count', 'Neutral Count'])
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            row_num = 2
            for batch in batches:
                ws.cell(row=row_num, column=1).value = batch.id
                ws.cell(row=row_num, column=2).value = batch.name
                ws.cell(row=row_num, column=3).value = batch.source.name
                ws.cell(row=row_num, column=4).value = batch.upload_date
                ws.cell(row=row_num, column=5).value = 'Yes' if batch.processed else 'No'
                ws.cell(row=row_num, column=6).value = batch.total_items
                
                if include_stats:
                    feedback_items = FeedbackItem.objects.filter(batch=batch)
                    avg_sentiment = feedback_items.aggregate(avg=Avg('sentiment_score'))['avg']
                    positive_count = feedback_items.filter(category__name='Positive').count()
                    negative_count = feedback_items.filter(category__name='Negative').count()
                    neutral_count = feedback_items.filter(category__name='Neutral').count()
                    
                    ws.cell(row=row_num, column=7).value = avg_sentiment
                    ws.cell(row=row_num, column=8).value = positive_count
                    ws.cell(row=row_num, column=9).value = negative_count
                    ws.cell(row=row_num, column=10).value = neutral_count
                
                row_num += 1
            
            # If include_items, add a sheet for each batch with its items
            if include_items:
                for batch in batches:
                    ws_items = wb.create_sheet(title=f"Batch {batch.id} Items")
                    
                    # Add header row
                    headers = ['ID', 'Content', 'Category', 'Rating', 'Sentiment Score', 'Processed']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws_items.cell(row=1, column=col_num)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')
                    
                    # Add data rows
                    row_num = 2
                    feedback_items = FeedbackItem.objects.filter(batch=batch)
                    for item in feedback_items:
                        ws_items.cell(row=row_num, column=1).value = item.id
                        ws_items.cell(row=row_num, column=2).value = item.content
                        ws_items.cell(row=row_num, column=3).value = item.category.name if item.category else None
                        ws_items.cell(row=row_num, column=4).value = item.rating
                        ws_items.cell(row=row_num, column=5).value = item.sentiment_score
                        ws_items.cell(row=row_num, column=6).value = 'Yes' if item.processed else 'No'
                        row_num += 1
            
            # Save to response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="batches_export.xlsx"'
            wb.save(response)
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            messages.warning(request, "Excel export is not available. Falling back to CSV format.")
            export_format = 'csv'
        
    if export_format == 'csv':  # Default or fallback
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="batches_export.csv"'
        
        writer = csv.writer(response)
        
        # Write header row
        headers = ['ID', 'Name', 'Source', 'Upload Date', 'Processed', 'Total Items']
        if include_stats:
            headers.extend(['Avg Sentiment', 'Positive Count', 'Negative Count', 'Neutral Count'])
        
        writer.writerow(headers)
        
        # Write data rows
        for batch in batches:
            row = [
                batch.id,
                batch.name,
                batch.source.name,
                batch.upload_date,
                'Yes' if batch.processed else 'No',
                batch.total_items
            ]
            
            if include_stats:
                feedback_items = FeedbackItem.objects.filter(batch=batch)
                avg_sentiment = feedback_items.aggregate(avg=Avg('sentiment_score'))['avg']
                positive_count = feedback_items.filter(category__name='Positive').count()
                negative_count = feedback_items.filter(category__name='Negative').count()
                neutral_count = feedback_items.filter(category__name='Neutral').count()
                
                row.extend([
                    avg_sentiment,
                    positive_count,
                    negative_count,
                    neutral_count
                ])
            
            writer.writerow(row)
            
            # If include_items, add rows for each feedback item
            if include_items:
                writer.writerow([])  # Empty row as separator
                writer.writerow(['Feedback Items for Batch: ' + batch.name])
                writer.writerow(['ID', 'Content', 'Category', 'Rating', 'Sentiment Score', 'Processed'])
                
                feedback_items = FeedbackItem.objects.filter(batch=batch)
                for item in feedback_items:
                    writer.writerow([
                        item.id,
                        item.content,
                        item.category.name if item.category else None,
                        item.rating,
                        item.sentiment_score,
                        'Yes' if item.processed else 'No'
                    ])
                
                writer.writerow([])  # Empty row as separator
    
    return response

@login_required
def export_feedback(request):
    """Export feedback items as CSV, Excel, or JSON"""
    import json
    import csv
    from django.db.models import Q
    
    # Get format from request
    export_format = request.GET.get('format', 'csv')
    
    # Get selection option
    selection = request.GET.get('selection', 'all')
    
    # Get selected fields
    fields = request.GET.getlist('fields', ['content', 'category', 'sentiment', 'rating', 'source', 'date'])
    
    # Determine which items to include
    if selection == 'all':
        items = FeedbackItem.objects.filter(client=request.user)
    elif selection == 'filtered':
        # Get filter parameters from request
        source = request.GET.get('source')
        category = request.GET.get('category')
        rating = request.GET.get('rating')
        date_range = request.GET.get('date_range')
        
        # Start with all items for this user
        items = FeedbackItem.objects.filter(client=request.user)
        
        # Apply filters
        if source:
            items = items.filter(source_id=source)
        if category:
            items = items.filter(category_id=category)
        if rating:
            items = items.filter(rating=rating)
        
        # Date range filtering
        if date_range == 'today':
            today = timezone.now().date()
            items = items.filter(created_at__date=today)
        elif date_range == 'week':
            start_of_week = timezone.now().date() - timedelta(days=timezone.now().weekday())
            items = items.filter(created_at__date__gte=start_of_week)
        elif date_range == 'month':
            start_of_month = timezone.now().date().replace(day=1)
            items = items.filter(created_at__date__gte=start_of_month)
        elif date_range == 'custom':
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            if date_from:
                items = items.filter(created_at__date__gte=date_from)
            if date_to:
                items = items.filter(created_at__date__lte=date_to)
    else:
        items = FeedbackItem.objects.filter(client=request.user)
    
    # Prepare data based on format
    if export_format == 'json':
        # Create JSON data
        data = []
        
        for item in items:
            item_data = {}
            
            if 'content' in fields:
                item_data['content'] = item.content
            if 'category' in fields:
                item_data['category'] = item.category.name if item.category else None
            if 'sentiment' in fields:
                item_data['sentiment_score'] = item.sentiment_score
            if 'rating' in fields:
                item_data['rating'] = item.rating
            if 'source' in fields:
                item_data['source'] = item.source.name
            if 'date' in fields:
                if item.feedback_date:
                    item_data['date'] = item.feedback_date.isoformat()
                else:
                    item_data['date'] = item.created_at.isoformat()
            
            data.append(item_data)
        
        # Create response with JSON data
        response = HttpResponse(json.dumps(data, indent=4), content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="feedback_export.json"'
        
    elif export_format == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Feedback"
            
            # Add header row
            headers = []
            if 'content' in fields: headers.append('Content')
            if 'category' in fields: headers.append('Category')
            if 'sentiment' in fields: headers.append('Sentiment Score')
            if 'rating' in fields: headers.append('Rating')
            if 'source' in fields: headers.append('Source')
            if 'date' in fields: headers.append('Date')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data rows
            for row_num, item in enumerate(items, 2):
                col_num = 1
                
                if 'content' in fields:
                    ws.cell(row=row_num, column=col_num).value = item.content
                    col_num += 1
                
                if 'category' in fields:
                    ws.cell(row=row_num, column=col_num).value = item.category.name if item.category else None
                    col_num += 1
                
                if 'sentiment' in fields:
                    ws.cell(row=row_num, column=col_num).value = item.sentiment_score
                    col_num += 1
                
                if 'rating' in fields:
                    ws.cell(row=row_num, column=col_num).value = item.rating
                    col_num += 1
                
                if 'source' in fields:
                    ws.cell(row=row_num, column=col_num).value = item.source.name
                    col_num += 1
                
                if 'date' in fields:
                    if item.feedback_date:
                        ws.cell(row=row_num, column=col_num).value = item.feedback_date
                    else:
                        ws.cell(row=row_num, column=col_num).value = item.created_at
            
            # Save to response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="feedback_export.xlsx"'
            wb.save(response)
        except ImportError:
            # Fallback to CSV if openpyxl is not available
            messages.warning(request, "Excel export is not available. Falling back to CSV format.")
            export_format = 'csv'
        
    if export_format == 'csv':  # Default or fallback
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="feedback_export.csv"'
        
        writer = csv.writer(response)
        
        # Write header row
        headers = []
        if 'content' in fields: headers.append('Content')
        if 'category' in fields: headers.append('Category')
        if 'sentiment' in fields: headers.append('Sentiment Score')
        if 'rating' in fields: headers.append('Rating')
        if 'source' in fields: headers.append('Source')
        if 'date' in fields: headers.append('Date')
        
        writer.writerow(headers)
        
        # Write data rows
        for item in items:
            row = []
            
            if 'content' in fields:
                row.append(item.content)
            
            if 'category' in fields:
                row.append(item.category.name if item.category else None)
            
            if 'sentiment' in fields:
                row.append(item.sentiment_score)
            
            if 'rating' in fields:
                row.append(item.rating)
            
            if 'source' in fields:
                row.append(item.source.name)
            
            if 'date' in fields:
                if item.feedback_date:
                    row.append(item.feedback_date)
                else:
                    row.append(item.created_at)
            
            writer.writerow(row)
    
    return response